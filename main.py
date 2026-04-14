import json
import os
import sys
import traceback
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintOptions


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "input", "WIP.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
CONFIG_DIR = os.path.join(BASE_DIR, "config")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_json(filename: str) -> dict:
    path = os.path.join(CONFIG_DIR, filename)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]

    column_map = {
        "RO #": "RO Number",
        "RO#": "RO Number",
        "Customer": "Owner",
        "Cust": "Owner",
        "Vehicle": "Vehicle",
        "Tech": "Technician",
        "Body Tech": "Technician",
        "Phase": "Phase",
        "Current Phase": "Phase",
        "Hours": "RO Hours",
        "RO Hours": "RO Hours",
    }

    renamed = {}
    for col in df.columns:
        renamed[col] = column_map.get(col, col)

    return df.rename(columns=renamed)


def autosize_columns(ws, min_width=12, max_width=45):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
            except Exception:
                value = ""
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def set_print_layout(ws, landscape=True, fit_width=1, fit_height=0):
    ws.print_options = PrintOptions(gridLines=True)
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_view.showGridLines = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2


def style_header_row(ws, row_num: int, fill_color: str = "D9EAF7"):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=fill_color)

    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_table(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_timestamp(ws, row_num: int = None):
    if row_num is None:
        row_num = ws.max_row + 2
    ts_str = datetime.now().strftime("Generated %A %m/%d/%Y %I:%M %p")
    cell = ws.cell(row=row_num, column=1)
    cell.value = ts_str
    cell.font = Font(italic=True, bold=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row_num


def clean_input_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = normalize_columns(df)

    required = ["Technician", "Phase", "RO Hours"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise RuntimeError(
            "Missing required column(s): " + ", ".join(missing) +
            ". Expected at least Technician, Phase, and RO Hours."
        )

    for optional in ["RO Number", "Owner", "Vehicle"]:
        if optional not in df.columns:
            df[optional] = ""

    df["Technician"] = df["Technician"].apply(normalize_text)
    df["Phase"] = df["Phase"].apply(normalize_text)
    df["Owner"] = df["Owner"].apply(normalize_text)
    df["Vehicle"] = df["Vehicle"].apply(normalize_text)
    df["RO Number"] = df["RO Number"].apply(normalize_text)
    df["RO Hours"] = pd.to_numeric(df["RO Hours"], errors="coerce")

    df = df.dropna(subset=["RO Hours"])
    df = df[df["Technician"] != ""].copy()

    return df


def build_phase_lookup(shop_phase_rules: dict):
    lookup = {}
    for phase_name, rule in shop_phase_rules.items():
        if isinstance(rule, dict):
            remaining_pct = float(rule.get("remaining_pct", 0))
            sort_order = int(rule.get("sort_order", 999))
            counts_in_load = bool(rule.get("counts_in_load", True))
        else:
            remaining_pct = float(rule)
            sort_order = 999
            counts_in_load = True

        lookup[str(phase_name).strip()] = {
            "remaining_pct": remaining_pct,
            "sort_order": sort_order,
            "counts_in_load": counts_in_load,
        }
    return lookup


def build_output(df: pd.DataFrame, shop_name: str, tech_capacity: dict, phase_lookup: dict):
    def calc_remaining(row):
        phase = normalize_text(row["Phase"])
        hours = float(row["RO Hours"])
        rule = phase_lookup.get(phase, {"remaining_pct": 0.0, "counts_in_load": False, "sort_order": 999})
        if not rule["counts_in_load"]:
            return 0.0
        return round(hours * float(rule["remaining_pct"]), 1)

    def phase_order(phase_name: str):
        return phase_lookup.get(phase_name, {"sort_order": 999}).get("sort_order", 999)

    df = df.copy()
    df["Remaining Hours"] = df.apply(calc_remaining, axis=1)
    df["Phase Sort"] = df["Phase"].apply(phase_order)

    tech_summary = (
        df.groupby("Technician", dropna=False)["Remaining Hours"]
        .sum()
        .reset_index()
    )

    tech_summary["Capacity"] = tech_summary["Technician"].map(tech_capacity)
    tech_summary = tech_summary.dropna(subset=["Capacity"]).copy()
    tech_summary["Capacity"] = pd.to_numeric(tech_summary["Capacity"], errors="coerce")
    tech_summary = tech_summary.dropna(subset=["Capacity"])

    tech_summary["Load %"] = (
        tech_summary["Remaining Hours"] / tech_summary["Capacity"] * 100
    ).round(0)

    avg_job = (
        df[df["Remaining Hours"] > 0]
        .groupby("Technician")["RO Hours"]
        .mean()
        .round(1)
    )
    job_count = (
        df[df["Remaining Hours"] > 0]
        .groupby("Technician")["RO Number"]
        .count()
    )

    tech_summary["Avg Job Size"] = tech_summary["Technician"].map(avg_job).fillna(0).round(1)
    tech_summary["Active Jobs"] = tech_summary["Technician"].map(job_count).fillna(0).astype(int)
    tech_summary = tech_summary.sort_values(["Load %", "Remaining Hours", "Technician"], ascending=[True, True, True]).reset_index(drop=True)
    tech_summary.insert(0, "Rank", range(1, len(tech_summary) + 1))

    timestamp = datetime.now().strftime("%m%d_%H%M")
    outfile = os.path.join(OUTPUT_DIR, f"WIP_Workload_{shop_name.replace(' ', '_')}_{timestamp}.xlsx")

    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        # Summary sheet
        summary_cols = ["Rank", "Technician", "Capacity", "Remaining Hours", "Load %", "Avg Job Size", "Active Jobs"]
        tech_summary.to_excel(writer, index=False, sheet_name="Summary", startrow=2)
        ws_sum = writer.sheets["Summary"]

        ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_cols))
        title = ws_sum.cell(row=1, column=1)
        title.value = f"{shop_name} Technician Workload Ranking"
        title.font = Font(bold=True, size=16)
        title.alignment = Alignment(horizontal="center", vertical="center")

        style_header_row(ws_sum, 3)
        style_data_table(ws_sum, 4, ws_sum.max_row, 1, len(summary_cols))
        set_print_layout(ws_sum, landscape=True, fit_width=1, fit_height=0)
        ws_sum.freeze_panes = "A4"

        for row in range(4, ws_sum.max_row + 1):
            ws_sum.cell(row=row, column=3).number_format = '0.0'
            ws_sum.cell(row=row, column=4).number_format = '0.0'
            ws_sum.cell(row=row, column=5).number_format = '0'
            ws_sum.cell(row=row, column=6).number_format = '0.0'

        add_timestamp(ws_sum, ws_sum.max_row + 2)
        autosize_columns(ws_sum)

        # Individual tech sheets
        detail_cols = ["RO Number", "Owner", "Vehicle", "Phase", "RO Hours", "Remaining Hours"]

        for tech in tech_summary["Technician"]:
            safe_name = tech[:31] if tech else "Unknown"
            ws = writer.book.create_sheet(title=safe_name)
            writer.sheets[safe_name] = ws

            df_t = df[df["Technician"] == tech].copy()
            df_t = df_t.sort_values(["Phase Sort", "RO Number", "Owner"])
            df_t = df_t[detail_cols]

            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(detail_cols))
            t = ws.cell(row=1, column=1)
            t.value = tech
            t.font = Font(bold=True, size=20)
            t.alignment = Alignment(horizontal="center", vertical="center")

            # Detail table
            for idx, col_name in enumerate(detail_cols, start=1):
                ws.cell(row=3, column=idx).value = col_name
            style_header_row(ws, 3)

            for r_idx, row in enumerate(df_t.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx).value = value

            if len(df_t) > 0:
                style_data_table(ws, 4, 3 + len(df_t), 1, len(detail_cols))
                for row in range(4, 4 + len(df_t)):
                    ws.cell(row=row, column=5).number_format = '0.0'
                    ws.cell(row=row, column=6).number_format = '0.0'

            # Tech snapshot section
            summary_start = max(6 + len(df_t), 8)
            ws.cell(row=summary_start, column=1).value = "Technician Snapshot"
            ws.cell(row=summary_start, column=1).font = Font(bold=True, size=12)

            tech_row = tech_summary[tech_summary["Technician"] == tech].iloc[0]
            snapshot_headers = ["Rank", "Capacity", "Remaining Hours", "Load %", "Avg Job Size", "Active Jobs"]
            snapshot_values = [
                int(tech_row["Rank"]),
                float(tech_row["Capacity"]),
                float(tech_row["Remaining Hours"]),
                float(tech_row["Load %"]),
                float(tech_row["Avg Job Size"]),
                int(tech_row["Active Jobs"]),
            ]

            for idx, header in enumerate(snapshot_headers, start=1):
                ws.cell(row=summary_start + 1, column=idx).value = header
                ws.cell(row=summary_start + 2, column=idx).value = snapshot_values[idx - 1]

            style_header_row(ws, summary_start + 1, fill_color="E8F4E8")
            style_data_table(ws, summary_start + 2, summary_start + 2, 1, len(snapshot_headers))
            for col in [2, 3, 5]:
                ws.cell(row=summary_start + 2, column=col).number_format = '0.0'
            ws.cell(row=summary_start + 2, column=4).number_format = '0'

            # Notes / definitions
            notes_row = summary_start + 5
            ws.cell(row=notes_row, column=1).value = "Definitions"
            ws.cell(row=notes_row, column=1).font = Font(bold=True, size=12)

            notes = [
                ("RO Hours", "Total labor hours from the WIP export."),
                ("Remaining Hours", "Labor remaining based on this shop's phase rules."),
                ("Load %", "Remaining Hours divided by weekly capacity."),
                ("Active Jobs", "Count of jobs with remaining hours greater than zero."),
            ]

            for idx, (label, desc) in enumerate(notes, start=1):
                ws.cell(row=notes_row + idx, column=1).value = label
                ws.cell(row=notes_row + idx, column=1).font = Font(bold=True)
                ws.cell(row=notes_row + idx, column=2).value = desc

            style_data_table(ws, notes_row + 1, notes_row + len(notes), 1, 2)

            add_timestamp(ws, notes_row + len(notes) + 2)
            ws.freeze_panes = "A4"
            set_print_layout(ws, landscape=True, fit_width=1, fit_height=0)
            autosize_columns(ws)

        # Remove default empty sheet if present and unused
        if "Sheet" in writer.book.sheetnames and writer.book["Sheet"].max_row == 1 and writer.book["Sheet"]["A1"].value is None:
            del writer.book["Sheet"]

    return outfile


def main():
    if not os.path.isfile(INPUT_FILE):
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}\n"
            f"Place your export in the input folder and name it WIP.xlsx"
        )

    shops_cfg = load_json("shops.json")
    phase_rules_cfg = load_json("phase_rules.json")
    tech_capacity_cfg = load_json("tech_capacity.json")

    shop_name = shops_cfg["default_shop"]
    if shop_name not in phase_rules_cfg:
        raise RuntimeError(f"No phase rules found for shop '{shop_name}'.")
    if shop_name not in tech_capacity_cfg:
        raise RuntimeError(f"No technician capacities found for shop '{shop_name}'.")

    df = pd.read_excel(INPUT_FILE)
    df = clean_input_dataframe(df)

    phase_lookup = build_phase_lookup(phase_rules_cfg[shop_name])
    outfile = build_output(
        df=df,
        shop_name=shop_name,
        tech_capacity=tech_capacity_cfg[shop_name],
        phase_lookup=phase_lookup,
    )

    print(f"Report generated: {outfile}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("An error occurred:")
        traceback.print_exc()
    finally:
        try:
            input("\nDone. Press Enter to close this window...")
        except Exception:
            pass
